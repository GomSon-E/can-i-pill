from __future__ import annotations

import argparse
import base64
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests


PROJECT_ROOT = Path(__file__).resolve().parent.parent
FILE_PATH_HEADER = "File Path"
VLM_TEXT_HEADER = "VLM Text"
VLM_RESULT_HEADER = "VLM Result"
TOKEN_PATTERN = re.compile(r"[0-9A-Za-z가-힣.%]+")
# 채점 규칙: 정답 토큰 집계 시 제외/정규화할 키
EXCLUDED_TOP_LEVEL_KEYS = {"처방전", "질환명", "질병분류기호"}
UNIT_STRIPPED_KEYS = {"1회투약량"}
PRESCRIPTION_PROMPT_IMAGE = (
    "이 처방전 이미지에서 처방의약품 관련 정보만 추출하세요.\n"
    "규칙:\n"
    "- 질환명, 질병분류기호, 각 의약품의 약품명, 1회투약량, 1일투여횟수, 총투약일수, 용법, 주의사항만 추출\n"
    "- 병원명, 주소, 전화번호, 환자 개인정보, 서명란, 약국사용란, 일반 안내문구는 제외\n"
    "- 약품별 정보는 한 덩어리로 묶어서 줄바꿈으로 구분\n"
    "- 약품명은 보이는 한글/영문/용량 표기를 가능한 그대로 유지\n"
    "- 설명, 요약, 해석, 마크다운, 코드블록, JSON 금지\n"
    "- 읽히지 않는 부분은 추측하지 말고 생략\n"
    "- 처방의약품 정보가 전혀 없으면 NONE 한 단어만 출력\n"
    "- 오직 추출된 텍스트만 출력\n"
    "\n"
    "출력 예시:\n"
    "질환명: 고혈압\n"
    "질병분류기호: I10\n"
    "약품명: 암로디핀정 5mg\n"
    "1회투약량: 1정\n"
    "1일투여횟수: 1\n"
    "총투약일수: 30\n"
    "용법: 아침 식후 30분\n"
    "주의사항: 발목 부종, 두통 발생 시 의사 상담\n"
    "\n"
    "약품명: 텔미사르탄정 40mg\n"
    "1회투약량: 1정\n"
    "1일투여횟수: 1\n"
    "총투약일수: 30\n"
    "용법: 아침 식후 30분"
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract prescription text from images via local VLM (Ollama) and "
            "write VLM Text / VLM Result into the existing prescription Excel file."
        )
    )
    parser.add_argument("--data-dir", default="data/처방전")
    parser.add_argument("--input", default="data/result/rx_results.xlsx")
    parser.add_argument(
        "--gt-json",
        default="data/처방전/text.json",
        help="Ground truth prescription JSON used for VLM scoring.",
    )
    parser.add_argument(
        "--vlm-model",
        default="qwen2.5vl:3b",
        help="Ollama VLM model name.",
    )
    parser.add_argument(
        "--ollama-url",
        default="http://127.0.0.1:11434/api/chat",
        help="Ollama chat API URL.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-extract for every row, overwriting existing VLM Text/Result.",
    )
    return parser.parse_args()


def resolve_project_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return PROJECT_ROOT / path


def normalize_prescription_name(name: str) -> str:
    return re.sub(r"[\s_]+", "", name).lower()


def load_answer_map(text_json_path: Path) -> dict[str, set[str]]:
    if not text_json_path.exists():
        return {}

    payload = json.loads(text_json_path.read_text(encoding="utf-8"))
    answer_map: dict[str, set[str]] = {}
    for item in payload:
        if not isinstance(item, dict):
            continue
        prescription_name = str(item.get("처방전", "")).strip()
        if not prescription_name:
            continue
        key = normalize_prescription_name(prescription_name)
        answer_map[key] = extract_answer_tokens(item)
    return answer_map


def strip_dose_unit(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    match = re.match(r"^([\d.]+)", text)
    if match:
        return match.group(1)
    return text


def extract_answer_tokens(value: Any) -> set[str]:
    tokens: set[str] = set()

    def walk(node: Any, depth: int = 0) -> None:
        if node is None:
            return
        if isinstance(node, dict):
            for key, child in node.items():
                if depth == 0 and key in EXCLUDED_TOP_LEVEL_KEYS:
                    continue
                if key in UNIT_STRIPPED_KEYS:
                    child = strip_dose_unit(child)
                walk(child, depth + 1)
            return
        if isinstance(node, list):
            for child in node:
                walk(child, depth + 1)
            return

        text = str(node).strip()
        if not text:
            return

        for token in TOKEN_PATTERN.findall(text):
            normalized = normalize_token(token)
            if normalized:
                tokens.add(normalized)

    walk(value)
    return tokens


def normalize_token(token: str) -> str:
    return re.sub(r"[^0-9a-z가-힣]+", "", token.lower())


def calculate_match_score(file_path: Path, extracted_text: str, answer_map: dict[str, set[str]]) -> str:
    key = normalize_prescription_name(file_path.stem)
    answer_tokens = answer_map.get(key)
    if not answer_tokens:
        return "N/A"

    extracted_tokens = {
        normalized
        for token in TOKEN_PATTERN.findall(extracted_text)
        for normalized in [normalize_token(token)]
        if normalized
    }
    matched_count = sum(1 for token in answer_tokens if token in extracted_tokens)
    total_count = len(answer_tokens)
    score = matched_count / total_count if total_count else 0
    return f"{score:.2%} ({matched_count}/{total_count})"


def encode_image_to_base64(image_path: Path) -> str:
    return base64.b64encode(image_path.read_bytes()).decode("utf-8")


def clean_output(text: str) -> str:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```[a-zA-Z0-9_-]*\s*\n?", "", cleaned)
        if cleaned.endswith("```"):
            cleaned = cleaned[: -len("```")]
        cleaned = cleaned.strip()
    return cleaned


def run_vlm_prescription(ollama_url: str, model_name: str, image_path: Path) -> str:
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": PRESCRIPTION_PROMPT_IMAGE,
                "images": [encode_image_to_base64(image_path)],
            }
        ],
        "stream": False,
        "options": {
            "temperature": 0,
            "top_p": 0.9,
            "top_k": 40,
            "num_predict": 2048,
        },
    }
    response = requests.post(ollama_url, json=payload, timeout=600)
    response.raise_for_status()
    data = response.json()
    message = data.get("message") or {}
    raw_content = str(message.get("content", "")).strip()
    if not raw_content:
        raise ValueError(f"Empty VLM response. Full payload: {data}")
    return clean_output(raw_content)


def read_header_map(ws) -> dict[str, int]:
    return {
        str(ws.cell(row=1, column=c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
    }


def ensure_vlm_headers(ws) -> dict[str, int]:
    """Add VLM Text / VLM Result columns at the end if missing. Never touches existing data."""
    header_map = read_header_map(ws)
    next_col = ws.max_column + 1
    for header in (VLM_TEXT_HEADER, VLM_RESULT_HEADER):
        if header not in header_map:
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
            next_col += 1
    return header_map


def cell_str(ws, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value)


def log(message: str) -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def main() -> int:
    args = parse_args()
    data_dir = resolve_project_path(args.data_dir)
    input_path = resolve_project_path(args.input)
    gt_json_path = resolve_project_path(args.gt_json)

    if not data_dir.exists():
        print(f"Data directory not found: {data_dir}", file=sys.stderr)
        return 1
    if not input_path.exists():
        print(f"Excel input not found: {input_path}", file=sys.stderr)
        return 1
    if not gt_json_path.exists():
        print(f"Ground truth JSON not found: {gt_json_path}", file=sys.stderr)
        return 1

    log(f"Loading ground truth JSON: {gt_json_path}")
    answer_map = load_answer_map(gt_json_path)
    log(f"Ground truth entries: {len(answer_map)}")

    log(f"Loading Excel: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    header_map = ensure_vlm_headers(ws)

    if FILE_PATH_HEADER not in header_map:
        print(f"Required header '{FILE_PATH_HEADER}' not found in sheet.", file=sys.stderr)
        return 1

    file_col = header_map[FILE_PATH_HEADER]
    vlm_text_col = header_map[VLM_TEXT_HEADER]
    vlm_result_col = header_map[VLM_RESULT_HEADER]

    def is_pending(row: int) -> bool:
        if args.force:
            return True
        text = cell_str(ws, row, vlm_text_col).strip()
        result = cell_str(ws, row, vlm_result_col).strip()
        if not text or not result:
            return True
        if text.startswith("ERROR:") or result == "ERROR":
            return True
        return False

    data_rows = [r for r in range(2, ws.max_row + 1) if cell_str(ws, r, file_col).strip()]
    pending_rows = [r for r in data_rows if is_pending(r)]
    log(
        f"Loaded {len(data_rows)} row(s). Pending: {len(pending_rows)}, "
        f"already filled: {len(data_rows) - len(pending_rows)}."
    )
    log(f"Using VLM model: {args.vlm_model} @ {args.ollama_url}")
    log(f"Started VLM export for {len(pending_rows)} image(s).")

    overall_start = time.monotonic()
    for image_idx, row in enumerate(pending_rows, start=1):
        image_path = Path(cell_str(ws, row, file_col)).resolve()
        file_name = image_path.name
        log(f"[VLM {image_idx}/{len(pending_rows)}] Starting image: {file_name}")
        image_start = time.monotonic()
        try:
            vlm_text = run_vlm_prescription(args.ollama_url, args.vlm_model, image_path)
            score = calculate_match_score(image_path, vlm_text, answer_map)
            ws.cell(row=row, column=vlm_text_col, value=vlm_text)
            ws.cell(row=row, column=vlm_result_col, value=score)
            elapsed = time.monotonic() - image_start
            log(
                f"[VLM {image_idx}/{len(pending_rows)}] Finished image: "
                f"{file_name} | VLM Result: {score} ({elapsed:.1f}s, {len(vlm_text)} chars)"
            )
        except Exception as exc:
            elapsed = time.monotonic() - image_start
            ws.cell(row=row, column=vlm_text_col, value=f"ERROR: {exc}")
            ws.cell(row=row, column=vlm_result_col, value="ERROR")
            log(
                f"[VLM {image_idx}/{len(pending_rows)}] Finished image: "
                f"{file_name} | ERROR after {elapsed:.1f}s: {exc}"
            )
        wb.save(input_path)
        log(f"[VLM {image_idx}/{len(pending_rows)}] Saved Excel through: {file_name}")

    total_elapsed = time.monotonic() - overall_start
    log(f"Saved Excel file: {input_path}")
    log(f"Completed VLM export in {total_elapsed:.1f}s.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
